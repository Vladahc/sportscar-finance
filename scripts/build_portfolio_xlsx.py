"""
Собирает Excel-таблицу портфеля: работа 20 тыс./мес + бот 100 тыс.,
20% прибыли бота и зарплата идут во вклад, облигации, акции, золото, биткоин.
Запуск: python scripts/build_portfolio_xlsx.py
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parents[1] / "docs" / "portfel-24m.xlsx"

YELLOW = PatternFill("solid", fgColor="FFF3C4")
HEAD = PatternFill("solid", fgColor="1A1A18")
OK = PatternFill("solid", fgColor="D8F3DC")
WARN = PatternFill("solid", fgColor="FDE8C8")
WHITE = Font(color="FFFFFF", bold=True, name="Calibri")
TITLE = Font(name="Calibri", size=16, bold=True)
H2 = Font(name="Calibri", size=12, bold=True)
BODY = Font(name="Calibri", size=11)
SMALL = Font(name="Calibri", size=10, italic=True, color="5C5C56")
THIN = Border(
    left=Side(style="thin", color="D4D4CE"),
    right=Side(style="thin", color="D4D4CE"),
    top=Side(style="thin", color="D4D4CE"),
    bottom=Side(style="thin", color="D4D4CE"),
)
WRAP = Alignment(wrap_text=True, vertical="center")


def money(cell):
    cell.number_format = '#,##0" ₽"'
    cell.font = BODY
    cell.border = THIN


def pct(cell):
    cell.number_format = "0.00%"
    cell.font = BODY
    cell.border = THIN


def inp(cell, value, kind="money"):
    cell.value = value
    cell.fill = YELLOW
    cell.font = BODY
    cell.border = THIN
    if kind == "money":
        cell.number_format = '#,##0" ₽"'
    elif kind == "pct":
        cell.number_format = "0.00%"
    elif kind == "int":
        cell.number_format = "0"


def label(ws, cell, text, font=BODY):
    ws[cell] = text
    ws[cell].font = font
    ws[cell].alignment = WRAP


def col_widths(ws, widths: dict[str, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_vvod(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Ввод"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C3"
    ws.row_dimensions[1].height = 28

    label(ws, "A1", "Портфель на 24 месяца: работа + бот на крипте", TITLE)
    label(
        ws,
        "A2",
        "Жёлтые ячейки можно менять. Лист «Месяцы» пересчитается сам. "
        "Цифры доходности — твой сценарий, не обещание рынка. Срез цен: лист «Примеры сегодня».",
        SMALL,
    )
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")

    label(ws, "A4", "Откуда деньги", H2)
    label(ws, "A5", "Откладываю с работы каждый месяц")
    inp(ws["C5"], 20_000)
    label(ws, "A6", "Старт бота на крипте")
    inp(ws["C6"], 100_000)
    label(ws, "A7", "Прибыль бота за месяц")
    inp(ws["C7"], 0.05, "pct")
    label(ws, "D7", "Смени, например, на 3% или 10% — итог справа обновится.", SMALL)
    label(ws, "A8", "Какую долю прибыли снимаем с бота")
    inp(ws["C8"], 0.20, "pct")
    label(ws, "A9", "Дата старта")
    ws["C9"] = datetime(2026, 9, 1)
    ws["C9"].fill = YELLOW
    ws["C9"].border = THIN
    ws["C9"].number_format = "YYYY-MM-DD"
    ws["C9"].font = BODY

    label(ws, "A11", "Куда кладём снятое с бота и 20 тысяч с работы (сумма долей = 100%)", H2)
    rows = [
        (12, "Спокойный вклад", 0.40),
        (13, "Облигации государства / фонды почти как вклад", 0.15),
        (14, "Акции компаний (лучше фонд на индекс, не одна бумага)", 0.20),
        (15, "Золото", 0.15),
        (16, "Биткоин (уже вне бота, без займа)", 0.10),
    ]
    for r, name, share in rows:
        label(ws, f"A{r}", name)
        inp(ws[f"C{r}"], share, "pct")
    label(ws, "A17", "Сумма долей (должно быть 100%)")
    ws["C17"] = "=C12+C13+C14+C15+C16"
    pct(ws["C17"])
    ws["D17"] = '=IF(ABS(C17-1)<0.0001,"норм","поправь доли: сумма не 100%")'
    ws["D17"].font = Font(name="Calibri", size=11, bold=True, color="B45309")

    label(ws, "A19", "Какой рост за год закладываем в расчёт (меняй под свой сценарий)", H2)
    years = [
        (20, "Вклад, % за год", 0.14),
        (21, "Облигации / фонды почти как вклад, % за год", 0.135),
        (22, "Акции компаний, % за год", 0.12),
        (23, "Золото, % за год", 0.08),
        (24, "Биткоин вне бота, % за год", 0.30),
    ]
    for r, name, val in years:
        label(ws, f"A{r}", name)
        inp(ws[f"C{r}"], val, "pct")
        label(ws, f"E{r}", "за месяц")
        ws[f"F{r}"] = f"=(1+C{r})^(1/12)-1"
        pct(ws[f"F{r}"])

    label(ws, "A26", "Сколько нужно на машину (из проекта, запас 10%)", H2)
    label(ws, "A27", "Вторичка с российским паспортом")
    inp(ws["C27"], 3_990_000)
    label(ws, "A28", "Салон, новая базовая")
    inp(ws["C28"], 5_720_000)

    label(ws, "A30", "Итог через 24 месяца", H2)
    results = [
        (31, "На боте", "=Месяцы!F26"),
        (32, "Во вкладе", "=Месяцы!I26"),
        (33, "В облигациях / фондах почти как вклад", "=Месяцы!J26"),
        (34, "В акциях", "=Месяцы!K26"),
        (35, "В золоте", "=Месяцы!L26"),
        (36, "В биткоине (вне бота)", "=Месяцы!M26"),
        (37, "Портфель рядом (всё кроме бота)", "=C32+C33+C34+C35+C36"),
        (38, "Всего денег", "=C31+C37"),
    ]
    for r, name, formula in results:
        label(ws, f"A{r}", name, H2 if r in (37, 38) else BODY)
        ws[f"C{r}"] = formula
        money(ws[f"C{r}"])
        if r == 38:
            ws[f"C{r}"].font = Font(name="Calibri", size=14, bold=True)

    label(ws, "A40", "Хватает ли на машину?")
    ws["C40"] = '=IF(C38>=C28,"да, на салон",IF(C38>=C27,"да, на вторичку","нет"))'
    ws["C40"].font = Font(name="Calibri", size=12, bold=True)
    ws["C40"].border = THIN
    ws.conditional_formatting.add(
        "C40",
        FormulaRule(formula=['C40="да, на салон"'], fill=OK),
    )
    ws.conditional_formatting.add(
        "C40",
        FormulaRule(formula=['C40="да, на вторичку"'], fill=WARN),
    )

    label(
        ws,
        "A42",
        "5% в месяц у бота — уже очень сильный сценарий. 21% в месяц два года подряд в проекте считается нереалистичным. "
        "Смотри docs/fuchersy-bot.md.",
        SMALL,
    )
    ws.merge_cells("A42:F43")

    dv = DataValidation(type="decimal", operator="between", formula1="0", formula2="2")
    dv.error = "Поставь долю от 0% до 200%"
    dv.add("C7:C8")
    dv.add("C12:C16")
    dv.add("C20:C24")
    ws.add_data_validation(dv)

    col_widths(ws, {"A": 58, "B": 12, "C": 18, "D": 55, "E": 12, "F": 14})
    ws.print_title_rows = "1:2"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1


def build_months(wb: Workbook) -> None:
    ws = wb.create_sheet("Месяцы")
    headers = [
        "Месяц",
        "Дата",
        "Бот на начало",
        "Прибыль бота",
        "Сняли 20% прибыли",
        "Бот на конец",
        "С работы",
        "Всего в портфель рядом",
        "Вклад",
        "Облигации / фонды",
        "Акции",
        "Золото",
        "Биткоин вне бота",
        "Портфель рядом",
        "Всего",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = HEAD
        cell.font = WHITE
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = "A1:O26"

    # Месяц 0 — старт.
    ws["A2"] = 0
    ws["B2"] = "=Ввод!C9"
    ws["B2"].number_format = "YYYY-MM-DD"
    ws["C2"] = 0
    ws["D2"] = 0
    ws["E2"] = 0
    ws["F2"] = "=Ввод!C6"
    ws["G2"] = 0
    ws["H2"] = 0
    for col in ("I", "J", "K", "L", "M"):
        ws[f"{col}2"] = 0
    ws["N2"] = "=I2+J2+K2+L2+M2"
    ws["O2"] = "=F2+N2"

    rates = {"I": "$F$20", "J": "$F$21", "K": "$F$22", "L": "$F$23", "M": "$F$24"}
    alloc = {"I": "$C$12", "J": "$C$13", "K": "$C$14", "L": "$C$15", "M": "$C$16"}

    for i in range(1, 25):
        r = i + 2
        p = r - 1
        ws[f"A{r}"] = i
        ws[f"B{r}"] = f"=EDATE($B$2,A{r})"
        ws[f"B{r}"].number_format = "YYYY-MM-DD"
        ws[f"C{r}"] = f"=F{p}"
        ws[f"D{r}"] = f"=C{r}*Ввод!$C$7"
        ws[f"E{r}"] = f"=IF(D{r}>0,D{r}*Ввод!$C$8,0)"
        ws[f"F{r}"] = f"=C{r}+D{r}-E{r}"
        ws[f"G{r}"] = "=Ввод!$C$5"
        ws[f"H{r}"] = f"=E{r}+G{r}"
        for col, rate in rates.items():
            # Сначала процент на уже лежащую сумму, потом новый взнос (в этом месяце ещё не растёт).
            ws[f"{col}{r}"] = f"={col}{p}*(1+Ввод!{rate})+H{r}*Ввод!{alloc[col]}"
        ws[f"N{r}"] = f"=I{r}+J{r}+K{r}+L{r}+M{r}"
        ws[f"O{r}"] = f"=F{r}+N{r}"

    for row in ws.iter_rows(min_row=2, max_row=26, min_col=1, max_col=15):
        for cell in row:
            cell.border = THIN
            cell.font = BODY
            if cell.column == 1:
                cell.number_format = "0"
            elif cell.column >= 3:
                cell.number_format = '#,##0" ₽"'

    for letter, w in zip("ABCDEFGHIJKLMNO", [10, 12, 16, 16, 18, 16, 14, 20, 14, 18, 14, 14, 18, 18, 16]):
        ws.column_dimensions[letter].width = w

    chart = LineChart()
    chart.title = "Всего денег по месяцам, ₽"
    chart.y_axis.title = "Рубли"
    chart.x_axis.title = "Месяц"
    chart.height = 10
    chart.width = 18
    chart.style = 10
    data = Reference(ws, min_col=15, min_row=1, max_row=26)
    cats = Reference(ws, min_col=1, min_row=2, max_row=26)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, "A28")

    chart2 = LineChart()
    chart2.title = "Бот и портфель рядом, ₽"
    chart2.y_axis.title = "Рубли"
    chart2.height = 10
    chart2.width = 18
    chart2.style = 12
    data2 = Reference(ws, min_col=6, min_row=1, max_col=6, max_row=26)
    data3 = Reference(ws, min_col=14, min_row=1, max_col=14, max_row=26)
    chart2.add_data(data2, titles_from_data=True)
    chart2.add_data(data3, titles_from_data=True)
    chart2.set_categories(cats)
    ws.add_chart(chart2, "I28")

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_title_rows = "1:1"


def build_help(wb: Workbook) -> None:
    ws = wb.create_sheet("Пояснение")
    ws.sheet_view.showGridLines = False
    label(ws, "A1", "Как читать таблицу", TITLE)
    lines = [
        "",
        "Жёлтые ячейки на листе «Ввод» — это ручки. Поменял прибыль бота — все 24 месяца и итог пересчитались.",
        "",
        "Бот. На старте 100 000 ₽ целиком на торговом счёте. Прибыль считается от суммы на начало месяца. "
        "Если прибыль плюс — снимаем долю (обычно 20%). Остальное остаётся на боте, объём растёт. "
        "Если месяц в минусе — с бота ничего не снимаем, счёт падает. 20 тысяч с работы всё равно кладём в портфель рядом.",
        "",
        "Портфель рядом. Сюда идут снятие с бота плюс отложенное с работы. Делим по долям. "
        "Сначала на уже лежащую сумму капает её процент за месяц, потом приходит новый взнос (в этом месяце он ещё не растёт).",
        "",
        "Почему такие доли по умолчанию. Вклад 40% — якорь. Облигации государства / фонды почти как вклад 15% — чуть гибче вклада. "
        "Акции 20% — рост, но прыгают. Золото 15% — страховка, если рубль слабеет. Биткоин 10% — сильный хвост, сильные просадки. "
        "Сайты с огромными процентами и займы незнакомцам сюда не кладём.",
        "",
        "Акции лучше фондом на индекс Мосбиржи (один пай), а не пятью случайными бумагами. Конкретные тикеры и цены — лист «Примеры сегодня».",
        "",
        "Это учебный расчёт при ровной прибыли, которую ты сам ставишь. Это не прогноз бота и не совет купить бумагу.",
        "",
        "Связанные файлы проекта: docs/analiz-xiaomi-su7-24m.md, docs/fuchersy-bot.md, docs/obzor.html.",
    ]
    for i, line in enumerate(lines, 2):
        label(ws, f"A{i}", line, SMALL if not line else BODY)
        ws.merge_cells(f"A{i}:F{i}")
        ws.row_dimensions[i].height = 36 if line else 10
    col_widths(ws, {"A": 120})


def build_examples(wb: Workbook) -> None:
    ws = wb.create_sheet("Примеры сегодня")
    ws.sheet_view.showGridLines = False
    label(ws, "A1", "Если решение принимаешь сегодня: цены и примеры", TITLE)
    ws.merge_cells("A1:G1")
    label(
        ws,
        "A2",
        "Срез 1 сентября 2026. Это ориентиры для первой покупки, не приказ брокеру. "
        "Перед сделкой открой стакан на Мосбирже: цена за день меняется. Это не индивидуальная рекомендация.",
        SMALL,
    )
    ws.merge_cells("A2:G2")

    label(ws, "A4", "Рынок вокруг копилки", H2)
    macro = [
        ("Дата среза", "1 сентября 2026", "Календарь проекта"),
        ("Ключевая ставка Банка России", "14%", "С 27 июля 2026, заседание 11 сентября 2026"),
        ("Доллар к рублю", "86,3793 ₽", "Банк России, 01.09.2026"),
        ("Юань к рублю", "12,8580 ₽", "Банк России, 01.09.2026"),
        ("Биткоин", "78 121 $ ≈ 6,78 млн ₽", "рынок, ориентир на дату сбора таблицы"),
        ("Золото, учётная цена Банка России", "12 671,47 ₽ за грамм", "с 1 сентября 2026, +97,15 ₽ к предыдущей"),
        ("Серебро, учётная цена Банка России", "195,12 ₽ за грамм", "с 1 сентября 2026"),
        ("Индекс Мосбиржи, ориентир конца августа", "около 2 115 пунктов", "неделя 24–28 августа 2026, −0,9%"),
        ("Вклады на 3–12 месяцев", "около 13–15%", "как в разборе проекта"),
        ("Фонды денежного рынка (LQDT, AKMM и похожие)", "около 14% годовых", "близко к ключевой ставке, не вклад"),
        ("ОФЗ, доходность к погашению, ориентир лета", "около 15%", "цена бумаги прыгает, если ставка изменится"),
    ]
    ws["A5"] = "Что"
    ws["B5"] = "Цифра"
    ws["C5"] = "Откуда / зачем"
    for col in ("A", "B", "C"):
        ws[f"{col}5"].fill = HEAD
        ws[f"{col}5"].font = WHITE
        ws[f"{col}5"].border = THIN
    for i, (a, b, c) in enumerate(macro, 6):
        ws[f"A{i}"] = a
        ws[f"B{i}"] = b
        ws[f"C{i}"] = c
        for col in ("A", "B", "C"):
            ws[f"{col}{i}"].border = THIN
            ws[f"{col}{i}"].font = BODY
            ws[f"{col}{i}"].alignment = WRAP

    label(ws, "A18", "Одна корзина, если решение принимаешь сегодня", H2)
    label(
        ws,
        "A19",
        "Первый месяц при настройках листа «Ввод»: с работы 20 000 ₽ + сняли с бота 1 000 ₽ = 21 000 ₽ в портфель рядом. "
        "Ниже — конкретная покупка на эти 21 000. Перед сделкой сверь цену в стакане.",
        BODY,
    )
    ws.merge_cells("A19:G19")
    ws.row_dimensions[19].height = 40
    basket_headers = ["Доля", "Куда", "Что купить сегодня", "Ориентир", "На сколько", "Что получится"]
    basket_rows = [
        [
            "40%",
            "Спокойный вклад",
            "Вклад 3–12 месяцев в крупном банке (в пределах страховки)",
            "13–15% годовых",
            "8 400 ₽",
            "Без лота: кладёшь всю сумму",
        ],
        [
            "15%",
            "Почти как вклад",
            "Паи денежного рынка: LQDT, AKMM или SBMM — если не хочешь открывать вклад каждый месяц",
            "около ключевой 14%",
            "3 150 ₽",
            "Паи на всю сумму 3 150 ₽",
        ],
        [
            "20%",
            "Акции",
            "Один фонд на индекс Мосбиржи: TMOS (запасной тот же смысл — SBMX или EQMX)",
            "TMOS 5,21 ₽ (27.08.2026)",
            "4 200 ₽",
            "около 800 паёв TMOS. Не 15 разных компаний руками",
        ],
        [
            "15%",
            "Золото",
            "Фонд золота: TGLD, GOLD, SBGD или AKGD. Не слиток",
            "учёт Банка России 12 671,47 ₽/г",
            "3 150 ₽",
            "паи на 3 150 ₽. Грамм слитка дороже всего этого куска",
        ],
        [
            "10%",
            "Биткоин вне бота",
            "Купить без займа и не трогать. Это не торговый счёт бота",
            "78 121 $ ≈ 6,78 млн ₽ за 1 BTC",
            "2 100 ₽",
            "≈ 0,00031 BTC",
        ],
    ]
    for col, h in enumerate(basket_headers, 1):
        cell = ws.cell(21, col, h)
        cell.fill = HEAD
        cell.font = WHITE
        cell.border = THIN
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[21].height = 28
    for ri, row in enumerate(basket_rows, 22):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, val)
            cell.border = THIN
            cell.font = BODY
            cell.alignment = WRAP
        ws.row_dimensions[ri].height = 40
    ws["A22"].fill = OK
    ws["A23"].fill = OK
    ws["A24"].fill = OK

    label(ws, "A28", "Как купить золото на эту сумму", H2)
    label(
        ws,
        "A29",
        "На лист «Ввод» в золото по умолчанию идёт 15% новых взносов. "
        "В первый месяц при боте 5% это примерно: прибыль бота 5 000 ₽, сняли 1 000 ₽, с работы 20 000 ₽, всего в портфель 21 000 ₽, в золото ≈ 3 150 ₽.",
        BODY,
    )
    ws.merge_cells("A29:G29")
    ws.row_dimensions[29].height = 48

    gold_headers = ["Как", "Тикер / что", "Ориентир цены", "На ≈ 3 150 ₽", "Плюс", "Минус"]
    gold_rows = [
        [
            "Проще всего",
            "Фонд золота на Мосбирже: TGLD, GOLD, SBGD, AKGD",
            "смотри пай в стакане",
            "паи на всю сумму 3 150 ₽",
            "не надо хранить слиток, можно продать в биржевой день",
            "комиссия фонда, цена зависит от металла и рубля",
        ],
        [
            "Слиток / монета",
            "золото 999 пробы",
            "12 671 ₽/г по Банку России",
            "на 3 150 ₽ даже 1 грамм не купить (нужно ≈ 12 700 ₽ на грамм)",
            "физический металл",
            "спред банка, хранение. Имеет смысл, когда в золоте уже десятки тысяч",
        ],
        [
            "10 грамм",
            "мерный слиток",
            "≈ 126 715 ₽ без спреда",
            "не в первый месяц",
            "понятная штука",
            "дорого относительно ежемесячного взноса",
        ],
    ]
    for col, h in enumerate(gold_headers, 1):
        cell = ws.cell(31, col, h)
        cell.fill = HEAD
        cell.font = WHITE
        cell.border = THIN
        cell.alignment = Alignment(wrap_text=True)
    for ri, row in enumerate(gold_rows, 32):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, val)
            cell.border = THIN
            cell.font = BODY
            cell.alignment = WRAP
        ws.row_dimensions[ri].height = 48

    label(ws, "A36", "Акции компаний: если решаешь сегодня", H2)
    label(
        ws,
        "A37",
        "Главный совет таблицы: не собирать «пять любимых бумаг» на 4 000 ₽ в месяц. "
        "Купи один биржевой фонд на индекс Мосбиржи — внутри уже Сбер, Лукойл, нефть, банки. "
        "Ниже — если всё же хочешь видеть конкретные имена. Фонд за год к концу августа 2026 был в минусе примерно на 20% — акции прыгают.",
        BODY,
    )
    ws.merge_cells("A37:G37")
    ws.row_dimensions[37].height = 56

    eq_headers = [
        "Приоритет",
        "Что купить",
        "Тикер",
        "Ориентир цены",
        "Дата цены",
        "На ≈ 4 200 ₽ (20% первого взноса)",
        "Зачем в копилку на машину",
    ]
    eq_rows = [
        [
            "1. Лучше так",
            "Фонд «индекс Мосбиржи», дивиденды внутри пая",
            "TMOS",
            "5,21 ₽ за пай",
            "27.08.2026",
            "около 800 паёв",
            "Одна покупка, внутри много компаний. Расходы фонда около 0,8% в год.",
        ],
        [
            "1. Тот же смысл",
            "Фонд «топ российских акций»",
            "SBMX",
            "15,13 ₽ за пай",
            "20.08.2026",
            "около 277 паёв",
            "Крупный фонд. За год к этой дате пай был примерно −22%.",
        ],
        [
            "1. Тот же смысл",
            "Фонд «индекс Мосбиржи»",
            "EQMX",
            "115,60 ₽ за пай",
            "20.08.2026",
            "36 паёв (чуть останется сдача)",
            "Дороже пай — меньше штук. Смысл тот же: индекс, не ставка на одну фирму.",
        ],
        [
            "2. Если очень хочешь одну бумагу",
            "Сбербанк, обыкновенные",
            "SBER",
            "275,85 ₽",
            "01.09.2026",
            "15 акций ≈ 4 138 ₽",
            "Самая ликвидная акция. Банк зарабатывает при высокой ставке, но цена прыгает вместе с рынком.",
        ],
        [
            "2. Не вместо фонда, а понимание индекса",
            "Лукойл",
            "LKOH",
            "4 324 ₽",
            "конец августа 2026",
            "1 акция ≈ 4 324 ₽ (весь кусок акций уйдёт в одну бумагу)",
            "Большая нефтяная компания, часто в топе индекса. Одна акция дорогая: на месячный взнос едва хватает штуки.",
        ],
        [
            "Не база",
            "Не класть всё в одну нефтяную или одну «историю»",
            "—",
            "—",
            "—",
            "—",
            "Для цели «машина через 2 года» важнее не угадать тикер, а не сжечь копилку. Индекс спокойнее эго.",
        ],
    ]
    for col, h in enumerate(eq_headers, 1):
        cell = ws.cell(39, col, h)
        cell.fill = HEAD
        cell.font = WHITE
        cell.border = THIN
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[39].height = 32
    for ri, row in enumerate(eq_rows, 40):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, val)
            cell.border = THIN
            cell.font = BODY
            cell.alignment = WRAP
        ws.row_dimensions[ri].height = 52
    ws["A40"].fill = OK
    ws["A41"].fill = OK
    ws["A42"].fill = OK

    label(ws, "A47", "Биткоин вне бота (10% новых взносов ≈ 2 100 ₽ в первый месяц при 5% бота)", H2)
    btc_lines = [
        ["Где", "Цена среза", "На 2 100 ₽", "Как не спутать с ботом"],
        [
            "Крупный обменник / биржа, без займа",
            "78 121 $ ≈ 6,78 млн ₽ за 1 BTC",
            "это ≈ 0,00031 BTC — пыль, но правило то же: купил и не трогать плечом",
            "Бот торгует своими 100 тысячами. Этот кусок — «купил и лежит», как в плане Б 10 тысяч на старте.",
        ],
    ]
    for col, h in enumerate(btc_lines[0], 1):
        cell = ws.cell(48, col, h)
        cell.fill = HEAD
        cell.font = WHITE
        cell.border = THIN
    for ci, val in enumerate(btc_lines[1], 1):
        cell = ws.cell(49, ci, val)
        cell.border = THIN
        cell.font = BODY
        cell.alignment = WRAP
    ws.row_dimensions[49].height = 56

    label(ws, "A51", "Спокойный вклад и облигации (55% новых взносов ≈ 11 550 ₽ в первый месяц)", H2)
    calm = [
        ["Куда", "Ориентир", "На 11 550 ₽", "Зачем"],
        [
            "Вклад 3–12 месяцев в крупном банке (в пределах страховки)",
            "13–15% годовых при ключевой 14%",
            "вся сумма 11 550 ₽ на вклад, без лота",
            "Якорь копилки. В расчёте на листе «Ввод» стоит 14% за год — поменяй, когда откроешь вклад.",
        ],
        [
            "Фонд денежного рынка LQDT / AKMM / SBMM",
            "около ключевой ставки",
            "паи на 11 550 ₽, если не хочешь заводить вклад каждый месяц",
            "Деньги можно вывести биржевым днём. Это не вклад: нет той же страховки банка.",
        ],
        [
            "ОФЗ с погашением близко к 2028 (когда нужна машина)",
            "доходность к погашению летом около 15%",
            "смотри лот на Мосбирже (часто от 1 бумаги)",
            "Если ставка вырастет, цена бумаги может просесть до погашения. Для 24 месяцев часто проще вклад.",
        ],
    ]
    for col, h in enumerate(calm[0], 1):
        cell = ws.cell(52, col, h)
        cell.fill = HEAD
        cell.font = WHITE
        cell.border = THIN
    for ri, row in enumerate(calm[1:], 53):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, val)
            cell.border = THIN
            cell.font = BODY
            cell.alignment = WRAP
        ws.row_dimensions[ri].height = 48

    label(
        ws,
        "A57",
        "Источники цен: Банк России (ставка, доллар, юань, учётные цены металлов с 01.09.2026); "
        "CoinGecko (биткоин); VBR (SBER 01.09.2026); Cbonds / страницы фондов (TMOS 27.08.2026, SBMX и EQMX 20.08.2026); "
        "Инвестминт (LKOH); Финаммаркет (индекс, неделя 24–28.08.2026). Пересчёт доходности на листе «Ввод» ты делаешь сам.",
        SMALL,
    )
    ws.merge_cells("A57:G58")
    ws.row_dimensions[57].height = 40

    col_widths(
        ws,
        {"A": 28, "B": 42, "C": 18, "D": 22, "E": 16, "F": 36, "G": 42},
    )
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_options.horizontalCentered = True


def main() -> None:
    wb = Workbook()
    build_vvod(wb)
    build_months(wb)
    build_help(wb)
    build_examples(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Записано {OUT}")


if __name__ == "__main__":
    main()
